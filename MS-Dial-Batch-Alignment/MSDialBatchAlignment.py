"""
Name: Bryan Roberts
Date: 8/8/2019
Description: Takes different batches of excel sheets in the same folder as script from the same LC-MS run and combines 
features into a new excel sheet called "results.xlsx".  Features are aligned based on annotation name.  Steps for 
getting matching annotation names:
-process initial batch with ~200 samples and all MSMS samples
-reduce this data export and make mzrt from remaining ~3000 features
-mzrt name will be in the format name_adduct_mz_rt
-process remaining data in batches using new mz/rt and correcting retention times for each batch
-use python script to align data
Sources:
https://automatetheboringstuff.com/
http://prime.psc.riken.jp/Metabolomics_Software/MS-DIAL/
Excel File Format:
Standard exported output from MS-Dial
"""

import openpyxl
import os.path

# return list of excel documents in folder
def getExcelSheets():
    excelSheets = []
    for file in os.listdir():
        if file[-5:] == '.xlsx':
            if file[0] != '~':
                excelSheets.append(os.path.join(os.getcwd(), file))
    return excelSheets

# return current filename
def getFileName(excelSheets, index):
    split = excelSheets[index].split(os.path.sep)
    return split[-1]

# opens and returns workbook of current excel sheet
def openWorkBook(excelSheets, index):
    wb = openpyxl.load_workbook(excelSheets[index])
    return wb

# returns sheet of current workbook
def makeSheet(wb):
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    return sheet

# makes new sheet called results and returns sheet
def makeResultsWorkBook():
    wb = openpyxl.Workbook()
    sheet = wb.active

    wb.save('results.xlsx')
    return wb

# copys all of first sheet into results sheet
def initializeResults(studySheet, resultsSheet):
    for c in range(1, studySheet.max_column + 1):
        for r in range(1, studySheet.max_row + 1):
            resultsSheet.cell(row = r, column=c).value = studySheet.cell(row = r, column = c).value

# copys sample informations from studySheet to resultsSheet
def copySampleInformation(studySheet, resultsSheet):
    count = 0
    col = resultsSheet.max_column
    for c in range(29, studySheet.max_column + 1):
        count += 1
        for r in range(1, 6):  
            resultsSheet.cell(row = r, column = col + count).value = studySheet.cell(row = r, column = c).value
    return col + 1

# determines if name matches for features in study row and results row
def nameMatch(studySheet, resultsSheet, studyRow, resultsRow):
    if studySheet.cell(row = studyRow, column = 4).value == resultsSheet.cell(row = resultsRow, column = 4).value:
        return True
    return False

# if feature matches, combine into results sheet
def findMatch(studySheet, resultsSheet, colStart):
    for studyRow in range(6, studySheet.max_row + 1):
        studyColumn = 29
        for resultsRow in range(6, resultsSheet.max_row + 1):
            if nameMatch(studySheet, resultsSheet, studyRow, resultsRow):
                for c in range(colStart, resultsSheet.max_column + 1):
                    resultsSheet.cell(row = resultsRow, column = c).value = studySheet.cell(row = studyRow, column = studyColumn).value
                    studyColumn += 1
    
"""
Execute main program
"""

if __name__ == "__main__":

    # initialize excelSheets
    excelSheets = getExcelSheets()
    wb = [openWorkBook(getExcelSheets(), num) for num in range(len(excelSheets))]
    sheet = [makeSheet(wb[num]) for num in range(len(wb))]
    
    # initialize results sheet
    resultsWB = makeResultsWorkBook()
    resultsSheet = makeSheet(resultsWB)
    
    # combine information from all excel sheets into results sheet with features aligned
    for i in range(len(sheet)):
        if i == 0:
            initializeResults(sheet[i], resultsSheet)
            print(getFileName(excelSheets, i))
        else:
            print(getFileName(excelSheets, i))
            startColumn = copySampleInformation(sheet[i], resultsSheet)
            findMatch(sheet[i], resultsSheet, startColumn)
    resultsWB.save('results.xlsx')
    print("done") 